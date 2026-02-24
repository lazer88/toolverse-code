"""
PDF Table → Excel  ·  Backend API
FastAPI + tabula-py + openpyxl
"""

import os
import re
import time
import uuid
import json
import threading
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import tabula
from fastapi import FastAPI, UploadFile, File, Form, Request, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
import openpyxl

# ═══════════════════════════════════════════════════
#  Configuration
# ═══════════════════════════════════════════════════

UPLOAD_DIR = Path("./temp_uploads")
OUTPUT_DIR = Path("./temp_outputs")
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

RATE_LIMIT = 3            # max conversions per window
RATE_WINDOW = 3600        # window = 1 hour (seconds)
MAX_FILE_SIZE = 100 * 1024 * 1024   # 100 MB
FILE_TTL = 3600           # auto-delete temp files after 1 hour

# ═══════════════════════════════════════════════════
#  App Init
# ═══════════════════════════════════════════════════

app = FastAPI(title="TableExtract API", version="1.0.0")

# CORS — adjust origins for your domain in production
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],          # ← replace with your frontend domain
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ═══════════════════════════════════════════════════
#  Rate Limiter (in-memory, per-IP)
# ═══════════════════════════════════════════════════

_rate_lock = threading.Lock()
_rate_log: dict[str, list[float]] = defaultdict(list)


def _get_client_ip(request: Request) -> str:
    """Extract real client IP (supports reverse proxy)."""
    forwarded = request.headers.get("X-Forwarded-For")
    if forwarded:
        return forwarded.split(",")[0].strip()
    real_ip = request.headers.get("X-Real-IP")
    if real_ip:
        return real_ip.strip()
    return request.client.host if request.client else "unknown"


def check_rate_limit(ip: str) -> dict:
    """Check & enforce rate limit. Returns status dict."""
    now = time.time()
    with _rate_lock:
        # Purge expired entries
        _rate_log[ip] = [ts for ts in _rate_log[ip] if now - ts < RATE_WINDOW]
        log = _rate_log[ip]

        if len(log) >= RATE_LIMIT:
            oldest = log[0]
            reset_at = oldest + RATE_WINDOW
            wait_sec = int(reset_at - now)
            wait_min = max(1, (wait_sec + 59) // 60)
            reset_time = datetime.fromtimestamp(reset_at).strftime("%H:%M")
            return {
                "allowed": False,
                "remaining": 0,
                "wait_seconds": wait_sec,
                "wait_minutes": wait_min,
                "reset_time": reset_time,
            }
        return {
            "allowed": True,
            "remaining": RATE_LIMIT - len(log),
        }


def record_usage(ip: str):
    with _rate_lock:
        _rate_log[ip].append(time.time())


# ═══════════════════════════════════════════════════
#  PDF Table Extraction Engine
# ═══════════════════════════════════════════════════

def count_pages(pdf_path: str) -> int | None:
    try:
        with open(pdf_path, "rb") as f:
            return len(re.findall(rb"/Type\s*/Page(?!s)", f.read()))
    except Exception:
        return None


def extract_page(pdf_path: str, page: int) -> list[pd.DataFrame]:
    """Extract tables from a single page (lattice → stream fallback)."""
    for mode in ("lattice", "stream"):
        try:
            kw = {"lattice": True} if mode == "lattice" else {"stream": True, "guess": True}
            dfs = tabula.read_pdf(pdf_path, pages=str(page), multiple_tables=True, silent=True, **kw)
            valid = [d for d in dfs if not d.empty and d.shape[0] > 0]
            if valid:
                return valid
        except Exception:
            pass
    return []


def run_extraction(pdf_path: str, pages: str = "all", parallel: bool = True, workers: int = 4) -> tuple[list, dict, list]:
    """
    Main extraction entry point.
    Returns (tables_as_dicts, stats, logs)
    """
    t0 = time.time()
    logs = []

    def log(msg, level="info"):
        logs.append({"msg": msg, "level": level, "time": datetime.now().strftime("%H:%M:%S")})

    fname = os.path.basename(pdf_path)
    log(f"Analyzing {fname}…")

    total_pages = count_pages(pdf_path)

    # Determine page list
    if pages.strip().lower() == "all":
        if total_pages and total_pages > 0:
            page_list = list(range(1, total_pages + 1))
            log(f"Detected {total_pages} pages")
        else:
            log("Page count unknown — bulk mode", "warn")
            tables = []
            for mode in ("lattice", "stream"):
                try:
                    kw = {"lattice": True} if mode == "lattice" else {"stream": True, "guess": True}
                    dfs = tabula.read_pdf(pdf_path, pages="all", multiple_tables=True, silent=True, **kw)
                    tables = [d for d in dfs if not d.empty]
                    if tables:
                        break
                except Exception:
                    pass
            elapsed = round(time.time() - t0, 2)
            log(f"Found {len(tables)} tables in {elapsed}s", "ok")
            return _to_dicts(tables), {"tables": len(tables), "rows": sum(len(t) for t in tables), "time": elapsed, "pages": "?"}, logs
    else:
        page_list = []
        for part in str(pages).split(","):
            p = part.strip()
            if "-" in p:
                a, b = p.split("-", 1)
                page_list.extend(range(int(a), int(b) + 1))
            else:
                page_list.append(int(p))

    n = len(page_list)
    log(f"{n} pages · {'parallel' if parallel and n > 2 else 'sequential'} mode")

    result_map = {}

    if parallel and n > 2:
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futs = {pool.submit(extract_page, pdf_path, pg): pg for pg in page_list}
            for f in as_completed(futs):
                pg = futs[f]
                tbls = f.result()
                if tbls:
                    result_map[pg] = tbls
                    log(f"Page {pg}: {len(tbls)} table(s)", "ok")
    else:
        for pg in page_list:
            tbls = extract_page(pdf_path, pg)
            if tbls:
                result_map[pg] = tbls
                log(f"Page {pg}: {len(tbls)} table(s)", "ok")

    all_tables = []
    for pg in sorted(result_map):
        all_tables.extend(result_map[pg])

    elapsed = round(time.time() - t0, 2)
    stats = {"tables": len(all_tables), "rows": sum(len(t) for t in all_tables), "time": elapsed, "pages": n}
    log(f"Done — {len(all_tables)} tables, {stats['rows']} rows in {elapsed}s", "ok")

    return _to_dicts(all_tables), stats, logs


def _to_dicts(tables: list[pd.DataFrame]) -> list[dict]:
    """Convert DataFrames to serializable dicts with cols/rows."""
    result = []
    for df in tables:
        df = df.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)
        # Smart header detection
        if len(df) > 1:
            row0 = df.iloc[0]
            nonnull = [v for v in row0 if pd.notna(v)]
            if nonnull and all(isinstance(v, str) for v in nonnull):
                cols = list(row0)
                if len(set(str(c) for c in cols if pd.notna(c))) == len([c for c in cols if pd.notna(c)]):
                    df.columns = cols
                    df = df.iloc[1:].reset_index(drop=True)
        cols = [str(c) for c in df.columns]
        rows = []
        for _, row in df.iterrows():
            rows.append([str(v) if pd.notna(v) else "" for v in row])
        result.append({"cols": cols, "rows": rows, "total_rows": len(rows)})
    return result


# ═══════════════════════════════════════════════════
#  Excel Generation (merged sheet, deduplicated headers)
# ═══════════════════════════════════════════════════

def generate_excel(tables: list[dict], output_path: str):
    """Generate .xlsx with all tables merged into one sheet, duplicate headers removed."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Tables"

    # Find max column count
    max_cols = max((len(t["cols"]) for t in tables), default=0)
    if max_cols == 0:
        wb.save(output_path)
        return

    def pad(row):
        r = list(row)
        while len(r) < max_cols:
            r.append("")
        return r

    # Collect all rows, mark headers
    all_rows = []
    header_keys = set()
    for tbl in tables:
        hk = "\x00".join(pad(tbl["cols"]))
        header_keys.add(hk)
        all_rows.append({"cells": pad(tbl["cols"]), "is_header": True})
        for r in tbl["rows"]:
            all_rows.append({"cells": pad(r), "is_header": False})

    # Deduplicate: keep each unique header only on first occurrence
    seen_headers = set()
    for row_info in all_rows:
        key = "\x00".join(row_info["cells"])
        if key in header_keys:
            if key in seen_headers:
                continue
            seen_headers.add(key)
        ws.append(row_info["cells"])

    # Auto-width
    for col_idx in range(1, max_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 40)

    wb.save(output_path)


# ═══════════════════════════════════════════════════
#  Temp File Cleanup (background)
# ═══════════════════════════════════════════════════

def cleanup_old_files():
    """Remove temp files older than FILE_TTL seconds."""
    now = time.time()
    for d in (UPLOAD_DIR, OUTPUT_DIR):
        for f in d.iterdir():
            if f.is_file() and now - f.stat().st_mtime > FILE_TTL:
                try:
                    f.unlink()
                except Exception:
                    pass

def _start_cleanup_timer():
    cleanup_old_files()
    t = threading.Timer(300, _start_cleanup_timer)  # every 5 min
    t.daemon = True
    t.start()

_start_cleanup_timer()


# ═══════════════════════════════════════════════════
#  API Endpoints
# ═══════════════════════════════════════════════════

@app.get("/api/health")
async def health():
    return {"status": "ok", "version": "1.0.0"}


@app.get("/api/rate-limit")
async def get_rate_limit(request: Request):
    """Check current rate limit status for the caller's IP."""
    ip = _get_client_ip(request)
    status = check_rate_limit(ip)
    return JSONResponse(content={"rate_limit": RATE_LIMIT, "window_seconds": RATE_WINDOW, **status})


@app.post("/api/convert")
async def convert_pdf(
    request: Request,
    file: UploadFile = File(...),
    pages: str = Form("all"),
    parallel: bool = Form(True),
    workers: int = Form(4),
):
    """
    Upload a PDF, extract tables, return preview + download token.
    Rate limited: 3 per hour per IP.
    """
    ip = _get_client_ip(request)

    # ── Rate limit ──
    status = check_rate_limit(ip)
    if not status["allowed"]:
        return JSONResponse(
            status_code=429,
            content={
                "error": "rate_limit_exceeded",
                "message": f"Conversion limit reached ({RATE_LIMIT} per hour).",
                "wait_minutes": status["wait_minutes"],
                "reset_time": status["reset_time"],
                "remaining": 0,
            },
        )

    # ── Validate file ──
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are accepted.")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail=f"File too large. Max {MAX_FILE_SIZE // (1024*1024)} MB.")

    # ── Save temp file ──
    task_id = uuid.uuid4().hex[:12]
    pdf_path = UPLOAD_DIR / f"{task_id}.pdf"
    pdf_path.write_bytes(content)

    # ── Extract ──
    try:
        tables, stats, logs = run_extraction(
            str(pdf_path),
            pages=pages,
            parallel=parallel,
            workers=min(workers, 8),
        )
    except Exception as e:
        pdf_path.unlink(missing_ok=True)
        raise HTTPException(status_code=500, detail=f"Extraction failed: {str(e)}")

    # ── Generate Excel ──
    download_token = None
    if tables:
        xlsx_path = OUTPUT_DIR / f"{task_id}.xlsx"
        try:
            generate_excel(tables, str(xlsx_path))
            download_token = task_id
        except Exception as e:
            logs.append({"msg": f"Excel generation error: {e}", "level": "err", "time": datetime.now().strftime("%H:%M:%S")})

    # ── Record usage (only on success) ──
    record_usage(ip)

    # ── Build preview data ──
    previews = []
    for i, tbl in enumerate(tables[:3]):
        previews.append({
            "id": i + 1,
            "cols": tbl["cols"],
            "rows": tbl["rows"][:6],
            "total_rows": tbl["total_rows"],
        })

    # ── Rate info for response ──
    updated_status = check_rate_limit(ip)

    # ── Cleanup uploaded PDF ──
    pdf_path.unlink(missing_ok=True)

    return JSONResponse(content={
        "stats": stats,
        "logs": logs,
        "previews": previews,
        "extra_count": max(0, len(tables) - 3),
        "download_token": download_token,
        "rate": {
            "remaining": updated_status.get("remaining", 0),
            "limit": RATE_LIMIT,
        },
    })


@app.get("/api/download/{token}")
async def download_file(token: str):
    """Download the generated Excel file by token."""
    # Sanitize token
    if not re.match(r'^[a-f0-9]{12}$', token):
        raise HTTPException(status_code=400, detail="Invalid token.")

    xlsx_path = OUTPUT_DIR / f"{token}.xlsx"
    if not xlsx_path.exists():
        raise HTTPException(status_code=404, detail="File not found or expired.")

    return FileResponse(
        path=str(xlsx_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"table_extract_{token}.xlsx",
    )


# ═══════════════════════════════════════════════════
#  Run
# ═══════════════════════════════════════════════════

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
