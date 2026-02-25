"""
Vercel Serverless Function: PDF Table → Excel
Auto-routes to /api/pdf-convert
Handles convert (POST), rate-limit check (GET), download (GET ?action=download&token=xxx)
"""
from http.server import BaseHTTPRequestHandler
import json, os, re, time, uuid, io, math
from datetime import datetime
from collections import defaultdict
from urllib.parse import urlparse, parse_qs

RATE_LIMIT = 3
RATE_WINDOW = 3600

_rate_log = defaultdict(list)
_file_store = {}

def _get_ip(headers):
    forwarded = headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return headers.get("X-Real-IP", "unknown")

def _check_rate(ip):
    now = time.time()
    _rate_log[ip] = [ts for ts in _rate_log[ip] if now - ts < RATE_WINDOW]
    log = _rate_log[ip]
    if len(log) >= RATE_LIMIT:
        oldest = log[0]
        reset_at = oldest + RATE_WINDOW
        wait = int(reset_at - now)
        wait_min = max(1, (wait + 59) // 60)
        rt = datetime.fromtimestamp(reset_at)
        return {"allowed": False, "remaining": 0, "wait_minutes": wait_min,
                "reset_time": rt.strftime("%H:%M"), "rate_limit": RATE_LIMIT}
    return {"allowed": True, "remaining": RATE_LIMIT - len(log), "rate_limit": RATE_LIMIT}

def _record(ip):
    _rate_log[ip].append(time.time())

def extract_tables_from_pdf(pdf_bytes, pages_str="all"):
    import pdfplumber

    logs = []
    def log(msg, level="info"):
        logs.append({"msg": msg, "level": level, "time": datetime.now().strftime("%H:%M:%S")})

    t0 = time.time()
    pdf = pdfplumber.open(io.BytesIO(pdf_bytes))
    total = len(pdf.pages)
    log(f"PDF loaded — {total} page(s)")

    if pages_str.strip().lower() == "all":
        page_list = list(range(total))
    else:
        page_list = []
        for part in pages_str.split(","):
            p = part.strip()
            if "-" in p:
                a, b = p.split("-", 1)
                page_list.extend(range(int(a)-1, min(int(b), total)))
            else:
                idx = int(p) - 1
                if 0 <= idx < total:
                    page_list.append(idx)

    all_tables = []
    for pi in page_list:
        page = pdf.pages[pi]
        tables = page.extract_tables()
        if tables:
            for tbl in tables:
                rows = [[str(c) if c else "" for c in row] for row in tbl if any(c for c in row)]
                if len(rows) >= 2:
                    all_tables.append({"cols": rows[0], "rows": rows[1:], "total_rows": len(rows) - 1})
            log(f"Page {pi+1}: {len(tables)} table(s)", "ok")

    elapsed = round(time.time() - t0, 2)
    total_rows = sum(t["total_rows"] for t in all_tables)
    log(f"Done — {len(all_tables)} tables, {total_rows} rows in {elapsed}s", "ok")
    stats = {"tables": len(all_tables), "rows": total_rows, "pages": len(page_list), "time": elapsed}
    pdf.close()
    return all_tables, stats, logs

def generate_excel(tables):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "All Tables"

    max_cols = max((len(t["cols"]) for t in tables), default=0)
    if max_cols == 0:
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    def pad(row):
        r = list(row)
        while len(r) < max_cols: r.append("")
        return r

    all_rows = []
    header_keys = set()
    for tbl in tables:
        hk = "\x00".join(pad(tbl["cols"]))
        header_keys.add(hk)
        all_rows.append(pad(tbl["cols"]))
        for r in tbl["rows"]:
            all_rows.append(pad(r))

    seen = set()
    for row in all_rows:
        key = "\x00".join(row)
        if key in header_keys:
            if key in seen: continue
            seen.add(key)
        ws.append(row)

    for ci in range(1, max_cols + 1):
        ml = 0
        for row in ws.iter_rows(min_col=ci, max_col=ci):
            for cell in row:
                if cell.value: ml = max(ml, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(ci)].width = min(ml + 2, 40)

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def _cleanup():
    now = time.time()
    expired = [k for k, v in _file_store.items() if now - v["created"] > 3600]
    for k in expired: del _file_store[k]

def parse_multipart(body_bytes, content_type):
    boundary = None
    for part in content_type.split(";"):
        p = part.strip()
        if p.startswith("boundary="):
            boundary = p[9:].strip('"')
            break
    if not boundary: return None, {}

    boundary_bytes = ("--" + boundary).encode()
    parts = body_bytes.split(boundary_bytes)
    file_data = None; file_name = ""; fields = {}

    for part in parts:
        if not part or part in (b"--\r\n", b"--"): continue
        if b"\r\n\r\n" in part:
            header_section, body = part.split(b"\r\n\r\n", 1)
        elif b"\n\n" in part:
            header_section, body = part.split(b"\n\n", 1)
        else: continue
        if body.endswith(b"\r\n"): body = body[:-2]

        header_str = header_section.decode("utf-8", errors="replace")
        name = ""; filename = ""
        for line in header_str.split("\n"):
            line = line.strip()
            if "Content-Disposition" in line:
                for token in line.split(";"):
                    token = token.strip()
                    if token.startswith("name="): name = token[5:].strip('"')
                    elif token.startswith("filename="): filename = token[9:].strip('"')

        if filename: file_data = body; file_name = filename
        elif name: fields[name] = body.decode("utf-8", errors="replace").strip()

    return (file_data, file_name), fields


class handler(BaseHTTPRequestHandler):

    def do_GET(self):
        qs = parse_qs(urlparse(self.path).query)
        action = qs.get("action", ["rate-limit"])[0]
        ip = _get_ip(self.headers)

        if action == "download":
            token = qs.get("token", [""])[0]
            if not re.match(r'^[a-f0-9]{12}$', token):
                self._json(400, {"error": "Invalid token"}); return
            entry = _file_store.get(token)
            if not entry:
                self._json(404, {"error": "File not found or expired"}); return
            data = entry["data"]
            self.send_response(200); self._cors()
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="table_extract_{token}.xlsx"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers(); self.wfile.write(data)
        else:
            status = _check_rate(ip)
            self._json(200, {"rate_limit": RATE_LIMIT, "window_seconds": RATE_WINDOW, **status})

    def do_POST(self):
        ip = _get_ip(self.headers)
        status = _check_rate(ip)
        if not status["allowed"]:
            self._json(429, {"error": "rate_limit_exceeded",
                "message": f"Conversion limit reached ({RATE_LIMIT} per hour).",
                "wait_minutes": status["wait_minutes"], "reset_time": status["reset_time"], "remaining": 0})
            return

        try:
            length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(length) if length else b""
            ct = self.headers.get("Content-Type", "")

            if "multipart/form-data" not in ct:
                self._json(400, {"error": "Expected multipart/form-data"}); return

            file_info, fields = parse_multipart(body, ct)
            if not file_info or not file_info[0]:
                self._json(400, {"error": "No file uploaded"}); return

            pdf_bytes, filename = file_info
            pages = fields.get("pages", "all")

            if not filename.lower().endswith(".pdf"):
                self._json(400, {"error": "Only PDF files accepted"}); return
            if len(pdf_bytes) > 100 * 1024 * 1024:
                self._json(400, {"error": "File too large (max 100MB)"}); return

            tables, stats, logs = extract_tables_from_pdf(pdf_bytes, pages)

            download_token = None
            if tables:
                xlsx_data = generate_excel(tables)
                token = uuid.uuid4().hex[:12]
                _cleanup()
                _file_store[token] = {"data": xlsx_data, "created": time.time()}
                download_token = token
                logs.append({"msg": f"Excel generated — {len(xlsx_data)//1024} KB", "level": "ok", "time": datetime.now().strftime("%H:%M:%S")})
            else:
                logs.append({"msg": "No tables detected", "level": "err", "time": datetime.now().strftime("%H:%M:%S")})

            _record(ip)
            updated = _check_rate(ip)

            previews = []
            for i, tbl in enumerate(tables[:3]):
                previews.append({"id": i+1, "cols": tbl["cols"], "rows": tbl["rows"][:6], "total_rows": tbl["total_rows"]})

            self._json(200, {"stats": stats, "logs": logs, "previews": previews,
                "extra_count": max(0, len(tables) - 3), "download_token": download_token,
                "rate": {"remaining": updated.get("remaining", 0), "limit": RATE_LIMIT}})

        except Exception as e:
            self._json(500, {"error": str(e)})

    def do_OPTIONS(self):
        self.send_response(200); self._cors()
        self.send_header("Content-Length", "0"); self.end_headers()

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def _json(self, code, data):
        body = json.dumps(data).encode()
        self.send_response(code); self._cors()
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers(); self.wfile.write(body)
